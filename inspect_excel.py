import openpyxl

def inspect_excel(filename):
    wb = openpyxl.load_workbook(filename, data_only=True)
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        sheet = wb[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        print(f"Headers: {headers}")
        for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=4, values_only=True)):
            print(f"Row {i+2}: {row}")

if __name__ == "__main__":
    inspect_excel("20260213-2_artprov.xlsx")
