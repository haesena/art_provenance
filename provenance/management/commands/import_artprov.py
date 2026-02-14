import openpyxl
from django.core.management.base import BaseCommand
from provenance.models import (
    ArtType, ArtworkGroup, Medium, Person, InstitutionType, 
    Institution, Artwork, Source, ProvenanceEvent, ArtworkRelationship
)
from datetime import datetime

class Command(BaseCommand):
    help = 'Import art provenance data from Excel'

    def handle(self, *args, **options):
        filename = '20260213-2_artprov.xlsx'
        wb = openpyxl.load_workbook(filename, data_only=True)

        self.stdout.write(self.style.SUCCESS('Starting import...'))

        # 1. Dropdown Tables
        self.import_art_types(wb['ArtTypes'])
        self.import_artwork_groups(wb['ArtworkGroups'])
        self.import_institution_types(wb['InstitutionTypes'])
        self.import_mediums(wb['Medium'])

        # 2. Key Entities
        self.import_persons(wb['Persons'])
        self.import_institutions(wb['Institutions'])
        self.import_artworks(wb['Artworks'])
        self.import_sources(wb['Sources'])

        # 3. Events and Relationships
        self.import_provenance_events(wb['ProvenanceEvents'])
        self.import_artwork_relationships(wb['Artworks']) # From the same sheet

        self.stdout.write(self.style.SUCCESS('Import completed successfully!'))

    def import_art_types(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                ArtType.objects.get_or_create(name=row[0])

    def import_artwork_groups(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                ArtworkGroup.objects.get_or_create(name=row[0])

    def import_institution_types(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                InstitutionType.objects.get_or_create(name=row[0])

    def import_mediums(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                art_type = None
                if row[1]:
                    art_type, _ = ArtType.objects.get_or_create(name=row[1])
                Medium.objects.get_or_create(
                    name=row[0], 
                    defaults={'type': art_type}
                )

    def import_persons(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Headers: ['Person', 'Family Name', 'First Name', 'DOB', 'DOD', 'Biography']
            if row[1]: # Family Name
                dob = self._format_date(row[3])
                dod = self._format_date(row[4])
                Person.objects.get_or_create(
                    family_name=row[1],
                    first_name=row[2] or '',
                    defaults={
                        'birth_date': dob,
                        'death_date': dod,
                        'biography': row[5] or ''
                    }
                )

    def import_institutions(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Headers: ['Name', 'Type', 'Place', 'StartDate', 'EndDate']
            if row[0]: # Name
                inst_type = None
                if row[1]:
                    inst_type, _ = InstitutionType.objects.get_or_create(name=row[1])
                Institution.objects.get_or_create(
                    name=row[0],
                    defaults={
                        'type': inst_type,
                        'place': row[2] or '',
                        'start_date': self._format_date(row[3]),
                        'end_date': self._format_date(row[4])
                    }
                )

    def import_artworks(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Headers: ['ArtworkGroup', 'Name', 'PossibleDuplicate', 'Medium', 'Dimension']
            if row[1]: # Name
                medium = None
                if row[3]:
                    medium, _ = Medium.objects.get_or_create(name=row[3])
                
                artwork, created = Artwork.objects.get_or_create(
                    name=row[1],
                    defaults={
                        'dimension': row[4] or '',
                        'medium': medium
                    }
                )
                if row[0]: # ArtworkGroup
                    group, _ = ArtworkGroup.objects.get_or_create(name=row[0])
                    artwork.groups.add(group)

    def import_sources(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Headers: ['Source', 'Type', 'Link']
            if row[0]:
                Source.objects.get_or_create(
                    source=row[0],
                    defaults={
                        'type': row[1] or '',
                        'link': row[2] if row[2] and str(row[2]).startswith('http') else None
                    }
                )

    def import_provenance_events(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Headers: ['Werk', 'Gruppe', 'Event', 'EventSeqNr', 'date', 'person', 'Institute', 'Certainty', 'Notes', 'Source1', 'Source2', None]
            artwork_name = row[0]
            if not artwork_name:
                continue
            
            artwork = Artwork.objects.filter(name=artwork_name).first()
            if not artwork:
                self.stdout.write(self.style.WARNING(f"Artwork not found: {artwork_name}"))
                continue

            person = None
            if row[5]:
                person = Person.objects.filter(family_name=row[5].split(',')[0]).first() # Rough match
            
            institution = None
            if row[6]:
                institution = Institution.objects.filter(name=row[6]).first()

            event = ProvenanceEvent.objects.create(
                artwork=artwork,
                event_type=row[2] or 'Unknown',
                sequence_number=row[3] if row[3] is not None else 0,
                date=self._format_date(row[4]) or '',
                person=person,
                institution=institution,
                certainty=row[7] if row[7] in dict(ProvenanceEvent.CERTAINTY_CHOICES) else None,
                notes=row[8] or ''
            )

            if row[9]: # Source1
                s1, _ = Source.objects.get_or_create(source=row[9])
                event.sources.add(s1)
            if row[10]: # Source2
                s2, _ = Source.objects.get_or_create(source=row[10])
                event.sources.add(s2)

    def import_artwork_relationships(self, sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Headers: ['ArtworkGroup', 'Name', 'PossibleDuplicate', 'Medium', 'Dimension']
            if row[1] and row[2]: # Name and PossibleDuplicate
                source_artwork = Artwork.objects.filter(name=row[1]).first()
                target_artwork = Artwork.objects.filter(name=row[2]).first()
                if source_artwork and target_artwork:
                    ArtworkRelationship.objects.get_or_create(
                        source_artwork=source_artwork,
                        target_artwork=target_artwork,
                        type='possible_match'
                    )

    def _format_date(self, value):
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.strftime('%d.%m.%Y')
        val_str = str(value).strip()
        if not val_str:
            return None
            
        # Try YYYY-MM-DD HH:MM:SS
        try:
            dt = datetime.strptime(val_str, '%Y-%m-%d %H:%M:%S')
            return dt.strftime('%d.%m.%Y')
        except ValueError:
            pass
            
        # Try YYYY-MM-DD
        try:
            dt = datetime.strptime(val_str, '%Y-%m-%d')
            return dt.strftime('%d.%m.%Y')
        except ValueError:
            pass
            
        return val_str
