import openpyxl
from django.core.management.base import BaseCommand
from provenance.models import (
    Person, Institution, Source, Auction, AuctionPerson, Exhibition
)
from datetime import datetime

class Command(BaseCommand):
    help = 'Import auctions and exhibitions from Excel'

    def handle(self, *args, **options):
        filename = '20260213-2_artprov.xlsx'
        wb = openpyxl.load_workbook(filename, data_only=True)

        self.stdout.write(self.style.SUCCESS('Starting import of Auctions and Exhibitions...'))

        if 'Auctions' in wb.sheetnames:
            self.import_auctions(wb['Auctions'])
        else:
            self.stdout.write(self.style.WARNING('Sheet "Auctions" not found.'))

        if 'Exhibitions' in wb.sheetnames:
            self.import_exhibitions(wb['Exhibitions'])
        else:
            self.stdout.write(self.style.WARNING('Sheet "Exhibitions" not found.'))

        self.stdout.write(self.style.SUCCESS('Import completed successfully!'))

    def import_auctions(self, sheet):
        self.stdout.write('Importing Auctions...')
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Headers: ['Date', 'Institution', 'Name', 'Notes', 'Auctioneer', 'Expert', 'Source 1', 'Source 2']
            date_val, inst_name, name, notes, auctioneer_name, expert_name, s1_name, s2_name = row[:8]
            
            if not name and not inst_name:
                continue

            institution = None
            if inst_name:
                institution, _ = Institution.objects.get_or_create(name=inst_name)

            auction, created = Auction.objects.get_or_create(
                name=name or f"Auction at {inst_name}",
                date=self._format_date(date_val) or '',
                defaults={
                    'institution': institution,
                    'notes': notes or ''
                }
            )

            # Sources
            for s_name in [s1_name, s2_name]:
                if s_name:
                    source, _ = Source.objects.get_or_create(source=s_name)
                    auction.sources.add(source)

            # Persons
            if auctioneer_name:
                self._link_person(auction, auctioneer_name, 'auctioneer')
            if expert_name:
                self._link_person(auction, expert_name, 'expert')

    def import_exhibitions(self, sheet):
        self.stdout.write('Importing Exhibitions...')
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Headers: ['DateStart', 'DateEnd', 'Institution', 'Name', 'Notes', 'Source']
            d_start, d_end, inst_name, name, notes, s_name = row[:6]

            if not name and not inst_name:
                continue

            institution = None
            if inst_name:
                institution, _ = Institution.objects.get_or_create(name=inst_name)

            exhibition, created = Exhibition.objects.get_or_create(
                name=name or f"Exhibition at {inst_name}",
                date_start=self._format_date(d_start) or '',
                defaults={
                    'date_end': self._format_date(d_end) or '',
                    'institution': institution,
                    'notes': notes or ''
                }
            )

            if s_name:
                source, _ = Source.objects.get_or_create(source=s_name)
                exhibition.sources.add(source)

    def _link_person(self, auction, person_fullname, role):
        # person_fullname is expected to be "Family, First" or just "Family"
        parts = str(person_fullname).split(',')
        family_name = parts[0].strip()
        first_name = parts[1].strip() if len(parts) > 1 else ''

        person, _ = Person.objects.get_or_create(
            family_name=family_name,
            first_name=first_name
        )
        AuctionPerson.objects.get_or_create(
            auction=auction,
            person=person,
            role=role
        )

    def _format_date(self, value):
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.strftime('%d.%m.%Y')
        val_str = str(value).strip()
        if not val_str or val_str.lower() == 'none':
            return None
            
        # Try YYYY
        try:
            if len(val_str) == 4 and val_str.isdigit():
                return val_str
        except:
            pass

        # Try YYYY-MM-DD HH:MM:SS
        try:
            dt = datetime.strptime(val_str, '%Y-%m-%d %H:%M:%S')
            return dt.strftime('%d.%m.%Y')
        except ValueError:
            pass
            
        # Try YYYY-MM-DD
        try:
            dt = datetime.strptime(val_str, '%Y-%m-%d')
            return dt.strftime('%d.%m.%Y')
        except ValueError:
            pass
            
        return val_str
